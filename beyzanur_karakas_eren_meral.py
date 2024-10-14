import tkinter
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import openpyxl
import pandas as pd


window = Tk()
window.geometry("720x250")
window.title("tk")
window.config(bg="#dedede")

class FileTypeNotSupported(BaseException):
    def init(self, message="YOU CHOOSE THE WRONG FILE TYPE. PLEASE MAKE A VALID SELECTION. PYTHON NOT SUPPORTED CSV"):
        self.message = message
        super().__init__(self.message)


listbox1 = tkinter.Listbox(selectmode=tkinter.EXTENDED, height=6, width=46)
listbox1.place(x=0, y=105)

listbox2 = tkinter.Listbox(selectmode=tkinter.EXTENDED, height=6, width=46)
listbox2.place(x=430, y=105)

def importListButton():
    filePath = tk.filedialog.askopenfilename() #hangi dosyayı seçtiğimizi kaydeder
    wb = load_workbook(filePath) #seçilen dosyayı açar
    ws = wb.active
    data = []
    for rows in range(2, 163):
        row_data = []
        for cols in range(1, 5):
            row_data.append(str(ws.cell(rows, cols).value))
        data.append(row_data)
    listbox1.delete(0, tk.END)
    for row in data:
        listbox1.insert(tk.END, " - ".join(row))

def addButton():
    selectedItem = listbox1.get(listbox1.curselection())
    listbox2.insert(END, selectedItem)
    listbox1.delete(listbox1.curselection())

def removeButton():
    selectedItem = listbox2.get(listbox2.curselection())
    listbox1.insert(END, selectedItem)
    listbox2.delete(listbox2.curselection())

def exportFile():
    selectedFileType = fileTypeCombobox.get()
    if (selectedFileType == "txt"):
        items = listbox2.get(0, tk.END)
        with open(f"ENGR 102 {enterWeekEntry.get()}.txt", "w") as file:
            weekEntry = enterWeekEntry.get()
            file.write(weekEntry + "\n")
            for item in items:
                file.write(item + "\n")
    elif (selectedFileType == "xls"):
        items = listbox2.get(0, tk.END)
        with open(f"ENGR 102 {enterWeekEntry.get()}.xls", "w") as file2:
            weekEntry = enterWeekEntry.get()
            file2.write(weekEntry + "\n")
            for item in items:
                file2.write(item + "\n")
    try:
        raise FileTypeNotSupported()
    except FileTypeNotSupported as e:
        if (selectedFileType == "csv"):
            print(e.message)

fileTypes = ["txt", "xls", "csv"]
fileTypeCombobox = ttk.Combobox(window, values=fileTypes, width=10)
fileTypeCombobox.set("txt")
fileTypeCombobox.place(x=185, y=210)

enterWeekEntry = Entry(font="Verdana 10", bg="white", fg="black", width=16)
enterWeekEntry.place(x=435, y=210)


class Interface():
    title = Label(text="AttendanceKeeper v1.0", bg="#dedede", fg="black", font="verdana 20 bold"    )
    title.place(x=190, y=0)

    selectStudentList = Label(text="Select student list Excel file : ", bg="#dedede", fg="black", font="verdana 13 bold")
    selectStudentList.grid(column=0, row=1, padx=0, pady=50)

    importListButton = Button(text="Import List", bg="#dedede", width="15", font="Verdana 9 bold", command=importListButton)
    importListButton.place(x=295, y=50)

    selectStudent = Label(text="Select a Student :", fg="black", bg="#dedede", font="verdana 13 bold")
    selectStudent.place(x=0, y=80)

    section = Label(text="Section :", fg="black", bg="#dedede", font="verdana 13 bold")
    section.place(x=315, y=80)

    addButton = Button(text="Add =>", bg="#dedede", width="16", height="2", font="Verdana 9  bold", command=addButton)
    addButton.place(x=287, y=130)

    removeButton = Button(text="<= Remove", bg="#dedede", width="16", height="2", font="Verdana 9 bold", command=removeButton)
    removeButton.place(x=287, y=170)

    attendedStudents = Label(text="Attended Students :", fg="black", bg="#dedede", font="verdana 13 bold")
    attendedStudents.place(x=435, y=80)

    selectFileType = Label(text="Please Select File Type :", fg="black", bg="#dedede", font="verdana 10 bold")
    selectFileType.place(x=0, y=210)

    enterWeek = Label(text="Please Enter Week :", fg="black", bg="#dedede", font="verdana 10 bold")
    enterWeek.place(x=285, y=210)

    exportFileButton = Button(text="Export as File", bg="#dedede", width="15", font="Verdana 9 bold", command=exportFile)
    exportFileButton.place(x=570, y=210)


wb = openpyxl.load_workbook('ENGR 102 Student List.xlsx') #excel dosyasını açar
ws = wb.active #çalışma sayfasını seçmemizi sağlar
data = [cell.value for cell in ws['D'][1:]] #hangi sütunda çalışacağımızı belirten kod
data = list(set(data)) #tekrara düşenleri silen kod
sectionCombobox = ttk.Combobox(window, values=data)
sectionCombobox.set("ENGR 102 01")
sectionCombobox.place(x=285, y=105)


class Filter():
    def filterlistbox(*args):
        # Excel dosyasından verileri okuma ve pandas veri çerçevesi oluşturma
        df = pd.read_excel("ENGR 102 Student List.xlsx")
        items = list(df["Section"])
        selectedItem = sectionCombobox.get()
        try:
            filtered_df = df.loc[df["Section"] == selectedItem]
            filteredList = filtered_df.values.tolist()
        except KeyError:
            filteredList = df.values.tolist()
        listbox1.delete(0, tk.END)
        for row in filteredList:
            row_str = " ".join(str(item) for item in row)
            listbox1.insert(tk.END, row_str)
    sectionCombobox.bind("<<ComboboxSelected>>", filterlistbox)


Interface()
Filter()
window.mainloop()